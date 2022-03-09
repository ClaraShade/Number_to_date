"""
This is a small script to extract a date from serial number of the product,
designed by my previous workplace.
The serial number must be in a form of 'YYMMDDXXX', where the last three digits corresponds to
the ID in the LOT.
It takes the .xlsx file with bunch of serial numbers and creates the .xlsx file with corresponding
dates
"""
from datetime import datetime
from openpyxl import load_workbook

wb1 = load_workbook('numbers.xlsx')
ws1 = wb1.active
for row_of_data in ws1.iter_rows():
    CHIP_ID = str(row_of_data[0].value)
    DATE_STRING = (CHIP_ID[0:6])
    dt = datetime.strptime(DATE_STRING, '%y%m%d')
    chip_date = dt.isoformat()
    row_of_data[0].value = chip_date
print("Ta dam!")
wb1.save('dates.xlsx')
idle = input("")
